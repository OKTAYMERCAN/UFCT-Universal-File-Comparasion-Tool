#!/usr/bin/env python3
"""
File Comparison Tool / Dosya Karşılaştırma Aracı  v4.0
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Usage / Kullanım:
  python dosya_karsilastir.py <folder_A> <folder_B>
  python dosya_karsilastir.py <folder_A> <folder_B> report.xlsx
  python dosya_karsilastir.py <folder_A> <folder_B> report.ods
  python dosya_karsilastir.py <folder_A> <folder_B> report.xlsx --mode quick
  python dosya_karsilastir.py <folder_A> <folder_B> report.ods  --mode full --algo sha256
  python dosya_karsilastir.py <folder_A> <folder_B> report.xlsx --lang tr
"""

import sys, subprocess, importlib, importlib.util, os

# ════════════════════════════════════════════════════════════════════════════
# 0.  AUTO-INSTALL
# ════════════════════════════════════════════════════════════════════════════

REQUIREMENTS = {"openpyxl": "openpyxl", "odf": "odfpy"}

def _check_and_install():
    missing = {m: p for m, p in REQUIREMENTS.items()
               if not importlib.util.find_spec(m)}
    if not missing:
        return

    # Language not yet selected — show bilingual prompt
    print()
    print("┌─────────────────────────────────────────────────────────────┐")
    print("│  Missing libraries detected / Eksik kütüphaneler tespit edildi │")
    print("└─────────────────────────────────────────────────────────────┘")
    for mod, pkg in missing.items():
        print(f"  • {mod:<12}  →  pip install {pkg}")
    print()

    while True:
        ans = input("  Auto-install? / Otomatik kurayım mı? [Y/E / N/H]: ").strip().upper()
        if ans in ("Y", "E", "YES", "EVET", ""):
            break
        if ans in ("N", "H", "NO", "HAYIR"):
            print("\n  Skipped / Atlandı. Manual install:")
            print("    pip install " + " ".join(missing.values()))
            sys.exit(1)
        print("  Y/E or N/H")

    print()
    for mod, pkg in missing.items():
        print(f"  Installing {pkg} ...", end="", flush=True)
        r = subprocess.run([sys.executable, "-m", "pip", "install", pkg, "--quiet"],
                           capture_output=True)
        if r.returncode == 0:
            print(" ✔")
        else:
            print(" ✘")
            print(r.stderr.decode(errors="replace"))
            sys.exit(1)

    print("\n  ✔ Done. Restarting / Tamamlandı. Yeniden başlatılıyor...\n")
    os.execv(sys.executable, [sys.executable] + sys.argv)

_check_and_install()

import hashlib, time, platform, argparse

# ════════════════════════════════════════════════════════════════════════════
# 1.  TRANSLATIONS
# ════════════════════════════════════════════════════════════════════════════

STRINGS = {
    "en": {
        # Language select
        "lang_header"       : "LANGUAGE",
        "lang_opt_en"       : "1) English  (default)",
        "lang_opt_tr"       : "2) Türkçe",
        "lang_prompt"       : "  Select / Seçin [1]: ",
        "lang_chosen_en"    : "  ✔ Language: English\n",
        "lang_chosen_tr"    : "  ✔ Dil: Türkçe\n",
        # Hardware
        "hw_header"         : "HARDWARE INFO",
        "hw_cpu"            : "CPU",
        "hw_os"             : "OS",
        "hw_python"         : "Python",
        "hw_aes_yes"        : "✔ Present — AES/SHA hardware acceleration",
        "hw_aes_no"         : "✘ Not present",
        "hw_sha_yes"        : "✔ Present — SHA256/SHA1 native instructions",
        "hw_sha_no"         : "✘ Not present",
        "hw_yes"            : "✔ Present",
        "hw_no"             : "✘ Not present",
        "bench_running"     : "Running benchmark on {mb} MB of data...",
        "bench_header"      : "  {n:<3} {algo:<12} {t:>8}   {spd:>10}   Bar",
        "bench_fastest"     : "  ← FASTEST",
        "bench_choose"      : "\n  Which algorithm would you like to use?",
        "bench_default"     : "    0) Default — fastest ({algo})",
        "bench_prompt"      : "\n  Your choice [0]: ",
        "bench_invalid"     : "  Invalid number.",
        "bench_chosen"      : "\n  ✔ Selected algorithm: {algo}\n",
        # Mode
        "mode_header"       : "COMPARISON MODE",
        "mode_1"            : "  1) ⚡ Quick Check   — File name & path list",
        "mode_1b"           : "                       No hash, instant result.",
        "mode_1c"           : "                       Which files are missing / extra?",
        "mode_2"            : "  2) 🔍 Full Check    — Includes hash comparison",
        "mode_2b"           : "                       Has file content changed?",
        "mode_2c"           : "                       (Opens benchmark + algorithm selector)",
        "mode_prompt"       : "  Your choice [1]: ",
        "mode_quick_ok"     : "\n  ✔ Mode: Quick Check (no hash)\n",
        "mode_full_ok"      : "\n  ✔ Mode: Full Check (hash enabled)\n",
        "mode_invalid"      : "  Enter 1 or 2.",
        # Scanning
        "scan_a"            : "[*] Scanning source A: {path}",
        "scan_b"            : "[*] Scanning source B: {path}",
        "scan_files"        : "    {n} file(s).",
        "scan_total"        : "[*] Total unique files: {n}",
        "scan_hashing"      : "[*] Calculating hashes ({algo})...",
        "scan_quick"        : "[*] Quick mode — skipping hash.",
        # Table headers
        "col_name_a"        : "File Name — A",
        "col_name_b"        : "File Name — B",
        "col_hash_a"        : "Hash A [{algo}]",
        "col_hash_b"        : "Hash B [{algo}]",
        "col_path_a"        : "Path — Source A",
        "col_path_b"        : "Path — Source B",
        "col_status"        : "Status",
        # Status values
        "st_equal"          : "✔ EQUAL",
        "st_diff"           : "✘ DIFFERENT",
        "st_both"           : "✔ IN BOTH",
        "st_only_a"         : "⚠ Only in A",
        "st_only_b"         : "⚠ Only in B",
        # Summary sheet
        "sum_title"         : "COMPARISON SUMMARY",
        "sum_src_a"         : "Source A",
        "sum_src_b"         : "Source B",
        "sum_format"        : "Format",
        "sum_mode"          : "Mode",
        "sum_mode_quick"    : "⚡ Quick (no hash)",
        "sum_mode_full"     : "🔍 Full — {algo}",
        "sum_total"         : "Total Files",
        "sum_both"          : "✔ Present in Both",
        "sum_hash_diff"     : "✘ Hash Different",
        "sum_only_a"        : "⚠ Only in A",
        "sum_only_b"        : "⚠ Only in B",
        # Result
        "res_saved"         : "  ✔  Report saved     : {path}",
        "res_format"        : "  Format             : {fmt}",
        "res_mode"          : "  Mode               : {mod}",
        "res_both"          : "  Present in both    : {n}",
        "res_diff"          : "  ✘ Hash different   : {n}",
        "res_only_a"        : "  ⚠ Only in A        : {n}",
        "res_only_b"        : "  ⚠ Only in B        : {n}",
        "res_mode_quick"    : "Quick (no hash)",
        "res_mode_full"     : "Full — {algo}",
        # Errors
        "err_folder"        : "ERROR: '{path}' is not a valid folder.",
        "err_ext"           : "ERROR: Unsupported extension '{ext}'. Use .xlsx or .ods.",
        "fmt_xlsx"          : "Excel (.xlsx)",
        "fmt_ods"           : "OpenDocument (.ods)",
    },
    "tr": {
        "lang_header"       : "DİL SEÇİMİ",
        "lang_opt_en"       : "1) English",
        "lang_opt_tr"       : "2) Türkçe  (varsayılan)",
        "lang_prompt"       : "  Seçin [2]: ",
        "lang_chosen_en"    : "  ✔ Language: English\n",
        "lang_chosen_tr"    : "  ✔ Dil: Türkçe\n",
        "hw_header"         : "DONANIM BİLGİSİ",
        "hw_cpu"            : "CPU",
        "hw_os"             : "İS",
        "hw_python"         : "Python",
        "hw_aes_yes"        : "✔ Var — AES/SHA donanım hızlandırma",
        "hw_aes_no"         : "✘ Yok",
        "hw_sha_yes"        : "✔ Var — SHA256/SHA1 native komutlar",
        "hw_sha_no"         : "✘ Yok",
        "hw_yes"            : "✔ Var",
        "hw_no"             : "✘ Yok",
        "bench_running"     : "{mb} MB veri üzerinde benchmark yapılıyor...",
        "bench_header"      : "  {n:<3} {algo:<12} {t:>8}   {spd:>10}   Bar",
        "bench_fastest"     : "  ← EN HIZLI",
        "bench_choose"      : "\n  Hangi algoritmayı kullanmak istersiniz?",
        "bench_default"     : "    0) Varsayılan — en hızlı ({algo})",
        "bench_prompt"      : "\n  Seçiminiz [0]: ",
        "bench_invalid"     : "  Geçersiz numara.",
        "bench_chosen"      : "\n  ✔ Seçilen algoritma: {algo}\n",
        "mode_header"       : "KARŞILAŞTIRMA MODU",
        "mode_1"            : "  1) ⚡ Hızlı Kontrol  — Dosya adı & konum listesi",
        "mode_1b"           : "                         Hash hesaplanmaz, anında sonuç.",
        "mode_1c"           : "                         Hangi dosyalar eksik / fazla?",
        "mode_2"            : "  2) 🔍 Tam Kontrol    — Hash hesaplaması dahil",
        "mode_2b"           : "                         Dosya içeriği değişmiş mi?",
        "mode_2c"           : "                         (Benchmark + algoritma seçimi açılır)",
        "mode_prompt"       : "  Seçiminiz [1]: ",
        "mode_quick_ok"     : "\n  ✔ Mod: Hızlı Kontrol (hash hesaplanmayacak)\n",
        "mode_full_ok"      : "\n  ✔ Mod: Tam Kontrol (hash hesaplanacak)\n",
        "mode_invalid"      : "  1 veya 2 girin.",
        "scan_a"            : "[*] Kaynak A taranıyor: {path}",
        "scan_b"            : "[*] Kaynak B taranıyor: {path}",
        "scan_files"        : "    {n} dosya.",
        "scan_total"        : "[*] Toplam benzersiz dosya: {n}",
        "scan_hashing"      : "[*] Hash hesaplanıyor ({algo})...",
        "scan_quick"        : "[*] Hızlı mod — hash hesaplanmıyor.",
        "col_name_a"        : "Dosya Adı — A",
        "col_name_b"        : "Dosya Adı — B",
        "col_hash_a"        : "Hash A [{algo}]",
        "col_hash_b"        : "Hash B [{algo}]",
        "col_path_a"        : "Konum — A Kaynağı",
        "col_path_b"        : "Konum — B Kaynağı",
        "col_status"        : "Durum",
        "st_equal"          : "✔ EŞİT",
        "st_diff"           : "✘ FARKLI",
        "st_both"           : "✔ İKİSİNDE DE VAR",
        "st_only_a"         : "⚠ Sadece A'da",
        "st_only_b"         : "⚠ Sadece B'de",
        "sum_title"         : "KARŞILAŞTIRMA ÖZETİ",
        "sum_src_a"         : "Kaynak A",
        "sum_src_b"         : "Kaynak B",
        "sum_format"        : "Format",
        "sum_mode"          : "Mod",
        "sum_mode_quick"    : "⚡ Hızlı (hash yok)",
        "sum_mode_full"     : "🔍 Tam — {algo}",
        "sum_total"         : "Toplam Dosya",
        "sum_both"          : "✔ İki Kaynakta Da Var",
        "sum_hash_diff"     : "✘ Hash Farklı",
        "sum_only_a"        : "⚠ Sadece A'da",
        "sum_only_b"        : "⚠ Sadece B'de",
        "res_saved"         : "  ✔  Rapor kaydedildi : {path}",
        "res_format"        : "  Format             : {fmt}",
        "res_mode"          : "  Mod                : {mod}",
        "res_both"          : "  İki kaynakta da var: {n}",
        "res_diff"          : "  ✘ Hash farklı      : {n}",
        "res_only_a"        : "  ⚠ Sadece A'da      : {n}",
        "res_only_b"        : "  ⚠ Sadece B'de      : {n}",
        "res_mode_quick"    : "Hızlı (hash yok)",
        "res_mode_full"     : "Tam — {algo}",
        "err_folder"        : "HATA: '{path}' geçerli bir klasör değil.",
        "err_ext"           : "HATA: Desteklenmeyen uzantı '{ext}'. .xlsx veya .ods kullanın.",
        "fmt_xlsx"          : "Excel (.xlsx)",
        "fmt_ods"           : "OpenDocument (.ods)",
    }
}

def t(lang, key, **kw):
    """Translate key in given language, format with kwargs."""
    return STRINGS[lang][key].format(**kw)

# ════════════════════════════════════════════════════════════════════════════
# 2.  LANGUAGE SELECTION
# ════════════════════════════════════════════════════════════════════════════

def lang_sec(forced=None):
    if forced in ("en", "tr"):
        return forced

    print("\n" + "=" * 62)
    print("  LANGUAGE / DİL")
    print("=" * 62)
    print()
    print("  1) English  (default / varsayılan)")
    print("  2) Türkçe")
    print()

    while True:
        try:
            secim = input("  Select / Seçin [1]: ").strip()
            if secim in ("", "1"):
                print("  ✔ Language: English\n")
                return "en"
            if secim == "2":
                print("  ✔ Dil: Türkçe\n")
                return "tr"
            print("  1 or 2 / 1 veya 2")
        except (EOFError, KeyboardInterrupt):
            return "en"

# ════════════════════════════════════════════════════════════════════════════
# 3.  HARDWARE & BENCHMARK
# ════════════════════════════════════════════════════════════════════════════

ALGORITHMS   = ["md5", "sha1", "sha256", "sha512", "blake2b", "blake2s"]
BENCHMARK_MB = 64

def _cpu_flags():
    try:
        with open("/proc/cpuinfo") as f:
            for line in f:
                if line.startswith("flags"):
                    return line.split(":", 1)[1].lower()
    except Exception:
        pass
    return ""

def _cpu_model():
    try:
        with open("/proc/cpuinfo") as f:
            for line in f:
                if "model name" in line:
                    return line.split(":", 1)[1].strip()
    except Exception:
        pass
    try:
        r = subprocess.run(["sysctl", "-n", "machdep.cpu.brand_string"],
                           capture_output=True, text=True)
        if r.returncode == 0:
            return r.stdout.strip()
    except Exception:
        pass
    try:
        import winreg
        key = winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE,
                             r"HARDWARE\DESCRIPTION\System\CentralProcessor\0")
        return winreg.QueryValueEx(key, "ProcessorNameString")[0].strip()
    except Exception:
        pass
    return platform.processor() or "Unknown"

def hw_info():
    flags = _cpu_flags()
    return {
        "cpu"    : _cpu_model(),
        "os"     : f"{platform.system()} {platform.release()}",
        "python" : platform.python_version(),
        "aes_ni" : "aes"    in flags,
        "sha_ni" : "sha_ni" in flags or "sha256_avx" in flags,
        "avx2"   : "avx2"   in flags,
        "sse4_2" : "sse4_2" in flags,
    }

def do_benchmark():
    data = os.urandom(BENCHMARK_MB * 1024 * 1024)
    res  = {}
    for algo in ALGORITHMS:
        try:
            t0 = time.perf_counter()
            h  = hashlib.new(algo); h.update(data); h.hexdigest()
            res[algo] = time.perf_counter() - t0
        except Exception:
            pass
    return res

def benchmark_and_choose(lang):
    hw = hw_info()
    print("\n" + "=" * 62)
    print(f"  {t(lang,'hw_header')}")
    print("=" * 62)
    print(f"  {t(lang,'hw_cpu'):<8}: {hw['cpu']}")
    print(f"  {t(lang,'hw_os'):<8}: {hw['os']}")
    print(f"  {t(lang,'hw_python'):<8}: {hw['python']}")
    print(f"  AES-NI : {t(lang,'hw_aes_yes') if hw['aes_ni'] else t(lang,'hw_aes_no')}")
    print(f"  SHA-NI : {t(lang,'hw_sha_yes') if hw['sha_ni'] else t(lang,'hw_sha_no')}")
    print(f"  AVX2   : {t(lang,'hw_yes') if hw['avx2']   else t(lang,'hw_no')}")
    print(f"  SSE4.2 : {t(lang,'hw_yes') if hw['sse4_2'] else t(lang,'hw_no')}")

    print(f"\n  {t(lang,'bench_running',mb=BENCHMARK_MB)}\n")
    results = do_benchmark()
    ranked  = sorted(results.items(), key=lambda x: x[1])
    max_spd = BENCHMARK_MB / ranked[0][1] if ranked else 1

    print("  " + "-" * 58)
    print(f"  {'#':<3} {'Algorithm':<12} {'Time':>8}   {'Speed':>10}   Bar")
    print("  " + "-" * 58)
    for idx, (algo, sec) in enumerate(ranked, 1):
        spd   = BENCHMARK_MB / sec
        bar   = "█" * int(spd / max_spd * 26)
        badge = t(lang, "bench_fastest") if idx == 1 else ""
        print(f"  {idx:<3} {algo:<12} {sec:>7.3f}s   {spd:>8.1f} MB/s  {bar}{badge}")
    print("  " + "-" * 58)

    print(t(lang, "bench_choose"))
    for idx, (algo, _) in enumerate(ranked, 1):
        print(f"    {idx}) {algo}")
    print(t(lang, "bench_default", algo=ranked[0][0]))

    while True:
        try:
            sel = input(t(lang, "bench_prompt")).strip()
            if sel in ("", "0"):
                chosen = ranked[0][0]; break
            n = int(sel)
            if 1 <= n <= len(ranked):
                chosen = ranked[n - 1][0]; break
            print(t(lang, "bench_invalid"))
        except (ValueError, KeyboardInterrupt):
            chosen = ranked[0][0]; break

    print(t(lang, "bench_chosen", algo=chosen.upper()))
    return chosen

# ════════════════════════════════════════════════════════════════════════════
# 4.  MODE SELECTION
# ════════════════════════════════════════════════════════════════════════════

def mode_select(lang):
    print("\n" + "=" * 62)
    print(f"  {t(lang,'mode_header')}")
    print("=" * 62)
    print()
    print(t(lang, "mode_1"))
    print(t(lang, "mode_1b"))
    print(t(lang, "mode_1c"))
    print()
    print(t(lang, "mode_2"))
    print(t(lang, "mode_2b"))
    print(t(lang, "mode_2c"))
    print()
    while True:
        try:
            sel = input(t(lang, "mode_prompt")).strip()
            if sel in ("", "1"):
                print(t(lang, "mode_quick_ok")); return "quick"
            if sel == "2":
                print(t(lang, "mode_full_ok"));  return "full"
            print(t(lang, "mode_invalid"))
        except (EOFError, KeyboardInterrupt):
            return "quick"

# ════════════════════════════════════════════════════════════════════════════
# 5.  FILE SCAN & HASH
# ════════════════════════════════════════════════════════════════════════════

def calc_hash(path, algo):
    h = hashlib.new(algo)
    try:
        with open(path, "rb") as f:
            for chunk in iter(lambda: f.read(65536), b""):
                h.update(chunk)
        return h.hexdigest()
    except (PermissionError, OSError):
        return "UNREADABLE"

def scan(folder):
    folder = os.path.abspath(folder)
    files  = {}
    for root, _, names in os.walk(folder):
        for name in names:
            full = os.path.join(root, name)
            rel  = os.path.relpath(full, folder)
            files[rel] = full
    return files

# ════════════════════════════════════════════════════════════════════════════
# 6a.  XLSX WRITER
# ════════════════════════════════════════════════════════════════════════════

def write_xlsx(src_a, src_b, out, mode, algo, da, db, all_files, lang):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.utils import get_column_letter

    H_BG="1F3864"; H_FG="FFFFFF"
    SHADE="EEF2FF"; WHITE="FFFFFF"
    GREEN="C6EFCE"; RED="FFC7CE"; YELLOW="FFEB9C"

    thin  = Side(style="thin", color="BDBDBD")
    bdr   = Border(left=thin, right=thin, top=thin, bottom=thin)

    def cell(ws, r, c, val, bg, fg="000000", bold=False, center=False):
        h = ws.cell(row=r, column=c, value=str(val) if val is not None else "")
        h.font      = Font(name="Arial", size=9, bold=bold, color=fg)
        h.fill      = PatternFill("solid", fgColor=bg)
        h.border    = bdr
        h.alignment = Alignment(vertical="center",
                                horizontal="center" if center else "left")

    wb = Workbook()
    ws = wb.active
    ws.title        = "Comparison" if lang == "en" else "Karşılaştırma"
    ws.freeze_panes = "A2"

    if mode == "quick":
        headers = [t(lang,"col_name_a"), t(lang,"col_name_b"),
                   t(lang,"col_path_a"), t(lang,"col_path_b"), t(lang,"col_status")]
        widths  = [36, 36, 52, 52, 24]
    else:
        headers = [t(lang,"col_name_a"), t(lang,"col_name_b"),
                   t(lang,"col_hash_a",algo=algo.upper()), t(lang,"col_hash_b",algo=algo.upper()),
                   t(lang,"col_path_a"), t(lang,"col_path_b"), t(lang,"col_status")]
        widths  = [36, 36, 34, 34, 52, 52, 24]

    for ci, (hdr, w) in enumerate(zip(headers, widths), 1):
        h = ws.cell(row=1, column=ci, value=hdr)
        h.font      = Font(name="Arial", size=10, bold=True, color=H_FG)
        h.fill      = PatternFill("solid", fgColor=H_BG)
        h.border    = bdr
        h.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 28

    eq=diff=only_a=only_b=0

    for ri, fname in enumerate(all_files, 2):
        va = fname in da; vb = fname in db
        pa = da.get(fname,""); pb = db.get(fname,"")
        zeb = SHADE if ri % 2 == 0 else WHITE

        if mode == "quick":
            if va and vb:   status=t(lang,"st_both");   sbg=GREEN;  eq+=1
            elif va:        status=t(lang,"st_only_a"); sbg=YELLOW; only_a+=1
            else:           status=t(lang,"st_only_b"); sbg=YELLOW; only_b+=1
            for ci,v in enumerate([fname if va else "",fname if vb else "",pa,pb],1):
                cell(ws,ri,ci,v,zeb)
            cell(ws,ri,5,status,sbg,bold=True,center=True)
        else:
            ha = calc_hash(pa,algo) if va else ""
            hb = calc_hash(pb,algo) if vb else ""
            if va and vb:
                if ha==hb: hbg=GREEN; status=t(lang,"st_equal"); sbg=GREEN;  eq+=1
                else:      hbg=RED;   status=t(lang,"st_diff");  sbg=RED;    diff+=1
            elif va:       hbg=YELLOW;status=t(lang,"st_only_a");sbg=YELLOW; only_a+=1
            else:          hbg=YELLOW;status=t(lang,"st_only_b");sbg=YELLOW; only_b+=1
            cell(ws,ri,1,fname if va else "",zeb)
            cell(ws,ri,2,fname if vb else "",zeb)
            cell(ws,ri,3,ha,hbg)
            cell(ws,ri,4,hb,hbg)
            cell(ws,ri,5,pa,zeb)
            cell(ws,ri,6,pb,zeb)
            cell(ws,ri,7,status,sbg,bold=True,center=True)

        ws.row_dimensions[ri].height = 15

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(all_files)+1}"

    if mode == "full":
        last = len(all_files)+1
        ws.conditional_formatting.add(f"C2:D{last}", FormulaRule(
            formula=["AND($C2<>\"\",$D2<>\"\",$C2=$D2)"],
            fill=PatternFill("solid", fgColor=GREEN)))
        ws.conditional_formatting.add(f"C2:D{last}", FormulaRule(
            formula=["AND($C2<>\"\",$D2<>\"\",$C2<>$D2)"],
            fill=PatternFill("solid", fgColor=RED)))

    # Summary sheet
    ws2 = wb.create_sheet("Summary" if lang=="en" else "Özet")
    ws2.column_dimensions["A"].width = 34
    ws2.column_dimensions["B"].width = 55

    thin2 = Side(style="thin", color="BDBDBD")
    bdr2  = Border(left=thin2,right=thin2,top=thin2,bottom=thin2)

    def oz(r, lbl, val, bg_l=SHADE, bg_v=WHITE, bold=False):
        for ci,(v,bg) in enumerate([(lbl,bg_l),(val,bg_v)],1):
            h = ws2.cell(row=r, column=ci, value=str(v))
            h.font      = Font(name="Arial",size=10,bold=bold,
                               color=H_FG if bg==H_BG else "000000")
            h.fill      = PatternFill("solid",fgColor=bg)
            h.border    = bdr2
            h.alignment = Alignment(vertical="center")
        ws2.row_dimensions[r].height=20

    mode_str = t(lang,"sum_mode_quick") if mode=="quick" else t(lang,"sum_mode_full",algo=algo.upper())
    oz(1, t(lang,"sum_title"),"",H_BG,H_BG,True)
    oz(2, t(lang,"sum_src_a"), os.path.abspath(src_a))
    oz(3, t(lang,"sum_src_b"), os.path.abspath(src_b))
    oz(4, t(lang,"sum_format"), t(lang,"fmt_xlsx"))
    oz(5, t(lang,"sum_mode"),   mode_str)
    oz(6, "","")
    oz(7, t(lang,"sum_total"),  len(all_files))
    oz(8, t(lang,"sum_both"),   eq,    SHADE, GREEN)
    row=9
    if mode == "full":
        oz(row, t(lang,"sum_hash_diff"), diff, SHADE, RED); row+=1
    oz(row,   t(lang,"sum_only_a"),  only_a, SHADE, YELLOW); row+=1
    oz(row,   t(lang,"sum_only_b"),  only_b, SHADE, YELLOW)

    wb.save(out)
    return eq, diff, only_a, only_b

# ════════════════════════════════════════════════════════════════════════════
# 6b.  ODS WRITER
# ════════════════════════════════════════════════════════════════════════════

def write_ods(src_a, src_b, out, mode, algo, da, db, all_files, lang):
    from odf.opendocument import OpenDocumentSpreadsheet
    from odf.style import (Style, TableColumnProperties, TableRowProperties,
                           TextProperties, TableCellProperties, ParagraphProperties)
    from odf.table import Table, TableColumn, TableRow, TableCell
    from odf.text  import P
    from odf.namespaces import FONS, STYLENS

    H_BG="#1F3864"; H_FG="#FFFFFF"; SH="#EEF2FF"; WH="#FFFFFF"
    GR="#C6EFCE"; RD="#FFC7CE"; YL="#FFEB9C"; BK="#000000"

    cache={}; doc=OpenDocumentSpreadsheet()

    def sty(name, bg=None, fg=BK, bold=False, align="left", sz="9pt"):
        if name in cache: return name
        cache[name]=True
        st=Style(name=name,family="table-cell")
        tp=TextProperties()
        tp.setAttrNS(FONS,"color",fg); tp.setAttrNS(FONS,"font-size",sz)
        tp.setAttrNS(FONS,"font-family","Arial")
        if bold: tp.setAttrNS(FONS,"font-weight","bold")
        st.addElement(tp)
        cp=TableCellProperties()
        if bg: cp.setAttrNS(FONS,"background-color",bg)
        for k in("border-left","border-right","border-top","border-bottom"):
            cp.setAttrNS(FONS,k,"0.5pt solid #BDBDBD")
        cp.setAttrNS(STYLENS,"vertical-align","middle")
        st.addElement(cp)
        pp=ParagraphProperties(); pp.setAttrNS(FONS,"text-align",align)
        st.addElement(pp); doc.automaticstyles.addElement(st); return name

    def cst(name,cm):
        st=Style(name=name,family="table-column")
        tcp=TableColumnProperties(); tcp.setAttrNS(STYLENS,"column-width",f"{cm}cm")
        st.addElement(tcp); doc.automaticstyles.addElement(st); return name

    def rst(name,cm):
        st=Style(name=name,family="table-row")
        trp=TableRowProperties()
        trp.setAttrNS(STYLENS,"row-height",f"{cm}cm")
        trp.setAttrNS(STYLENS,"use-optimal-row-height","false")
        st.addElement(trp); doc.automaticstyles.addElement(st); return name

    def cel(val,s):
        tc=TableCell(stylename=s)
        tc.addElement(P(text=str(val) if val is not None else ""))
        return tc

    sH =sty("sH",H_BG,H_FG,bold=True,align="center",sz="10pt")
    sA =sty("sA",SH,BK);  sW=sty("sW",WH,BK)
    sGR=sty("sGR",GR,BK); sRD=sty("sRD",RD,BK); sYL=sty("sYL",YL,BK)
    sDG=sty("sDG",GR,BK,bold=True,align="center")
    sDR=sty("sDR",RD,BK,bold=True,align="center")
    sDY=sty("sDY",YL,BK,bold=True,align="center")

    rH=rst("rH",0.9); rV=rst("rV",0.55)

    if mode=="quick":
        cols=[("q0",8.5),("q1",8.5),("q2",14.0),("q3",14.0),("q4",5.5)]
        hdrs=[t(lang,"col_name_a"),t(lang,"col_name_b"),
              t(lang,"col_path_a"),t(lang,"col_path_b"),t(lang,"col_status")]
    else:
        cols=[("q0",8.5),("q1",8.5),("q2",9.5),("q3",9.5),
              ("q4",13.5),("q5",13.5),("q6",5.5)]
        hdrs=[t(lang,"col_name_a"),t(lang,"col_name_b"),
              t(lang,"col_hash_a",algo=algo.upper()),t(lang,"col_hash_b",algo=algo.upper()),
              t(lang,"col_path_a"),t(lang,"col_path_b"),t(lang,"col_status")]

    for cn,cm in cols: cst(cn,cm)
    tbl=Table(name="Comparison" if lang=="en" else "Karşılaştırma")
    for cn,_ in cols: tbl.addElement(TableColumn(stylename=cn))

    trh=TableRow(stylename=rH)
    for h in hdrs: trh.addElement(cel(h,sH))
    tbl.addElement(trh)

    eq=diff=only_a=only_b=0

    for i,fname in enumerate(all_files):
        va=fname in da; vb=fname in db
        pa=da.get(fname,""); pb=db.get(fname,"")
        zb=sA if i%2==0 else sW

        if mode=="quick":
            if va and vb:   st2=t(lang,"st_both");   sd=sDG; eq+=1
            elif va:        st2=t(lang,"st_only_a"); sd=sDY; only_a+=1
            else:           st2=t(lang,"st_only_b"); sd=sDY; only_b+=1
            tr=TableRow(stylename=rV)
            for v,s in [(fname if va else "",zb),(fname if vb else "",zb),
                        (pa,zb),(pb,zb),(st2,sd)]:
                tr.addElement(cel(v,s))
            tbl.addElement(tr)
        else:
            ha=calc_hash(pa,algo) if va else ""
            hb=calc_hash(pb,algo) if vb else ""
            if va and vb:
                if ha==hb: hs=sGR;st2=t(lang,"st_equal"); sd=sDG; eq+=1
                else:      hs=sRD;st2=t(lang,"st_diff");  sd=sDR; diff+=1
            elif va:       hs=sYL;st2=t(lang,"st_only_a");sd=sDY; only_a+=1
            else:          hs=sYL;st2=t(lang,"st_only_b");sd=sDY; only_b+=1
            tr=TableRow(stylename=rV)
            for v,s in [(fname if va else "",zb),(fname if vb else "",zb),
                        (ha,hs),(hb,hs),(pa,zb),(pb,zb),(st2,sd)]:
                tr.addElement(cel(v,s))
            tbl.addElement(tr)

    doc.spreadsheet.addElement(tbl)

    # Summary
    sOH=sty("sOH",H_BG,H_FG,bold=True,sz="11pt")
    sOL=sty("sOL",SH, BK,bold=True,sz="10pt"); sOV=sty("sOV",WH,BK,sz="10pt")
    sOG=sty("sOG",GR, BK,sz="10pt"); sOR2=sty("sOR2",RD,BK,sz="10pt")
    sOY=sty("sOY",YL, BK,sz="10pt")
    cst("o0",8.0); cst("o1",16.0)

    toz=Table(name="Summary" if lang=="en" else "Özet")
    toz.addElement(TableColumn(stylename="o0"))
    toz.addElement(TableColumn(stylename="o1"))

    def oz(lbl,val,sl=sOL,sv=sOV):
        row=TableRow(stylename=rV)
        row.addElement(cel(lbl,sl)); row.addElement(cel(val,sv))
        toz.addElement(row)

    th2=TableRow(stylename=rH)
    th2.addElement(cel(t(lang,"sum_title"),sOH)); th2.addElement(cel("",sOH))
    toz.addElement(th2)

    mode_str=t(lang,"sum_mode_quick") if mode=="quick" else t(lang,"sum_mode_full",algo=algo.upper())
    oz(t(lang,"sum_src_a"), os.path.abspath(src_a))
    oz(t(lang,"sum_src_b"), os.path.abspath(src_b))
    oz(t(lang,"sum_format"),t(lang,"fmt_ods"))
    oz(t(lang,"sum_mode"),  mode_str)
    oz("","")
    oz(t(lang,"sum_total"), len(all_files))
    oz(t(lang,"sum_both"),  eq,    sOL,sOG)
    if mode=="full": oz(t(lang,"sum_hash_diff"),diff,sOL,sOR2)
    oz(t(lang,"sum_only_a"),only_a,sOL,sOY)
    oz(t(lang,"sum_only_b"),only_b,sOL,sOY)

    doc.spreadsheet.addElement(toz)
    doc.save(out)
    return eq, diff, only_a, only_b

# ════════════════════════════════════════════════════════════════════════════
# 7.  MAIN
# ════════════════════════════════════════════════════════════════════════════

def run(src_a, src_b, out, mode, algo, lang):
    ext = os.path.splitext(out)[1].lower()
    if ext not in (".xlsx", ".ods"):
        print(t(lang, "err_ext", ext=ext)); sys.exit(1)

    print(f"\n{t(lang,'scan_a',path=src_a)}")
    da = scan(src_a)
    print(t(lang,"scan_files",n=len(da)))
    print(t(lang,"scan_b",path=src_b))
    db = scan(src_b)
    print(t(lang,"scan_files",n=len(db)))
    all_files = sorted(set(da)|set(db))
    print(t(lang,"scan_total",n=len(all_files)))

    if mode=="full":
        print(t(lang,"scan_hashing",algo=algo.upper()))
    else:
        print(t(lang,"scan_quick"))

    if ext==".xlsx":
        eq,diff,oa,ob = write_xlsx(src_a,src_b,out,mode,algo,da,db,all_files,lang)
        fmt_label = t(lang,"fmt_xlsx")
    else:
        eq,diff,oa,ob = write_ods(src_a,src_b,out,mode,algo,da,db,all_files,lang)
        fmt_label = t(lang,"fmt_ods")

    mod_label = (t(lang,"res_mode_quick") if mode=="quick"
                 else t(lang,"res_mode_full",algo=algo.upper()))

    print(f"\n{'='*58}")
    print(t(lang,"res_saved", path=out))
    print(t(lang,"res_format",fmt=fmt_label))
    print(t(lang,"res_mode",  mod=mod_label))
    print(t(lang,"res_both",  n=eq))
    if mode=="full":
        print(t(lang,"res_diff",  n=diff))
    print(t(lang,"res_only_a",n=oa))
    print(t(lang,"res_only_b",n=ob))
    print(f"{'='*58}\n")

def main():
    parser = argparse.ArgumentParser(
        description="File comparison tool — xlsx/ods output, EN/TR language"
    )
    parser.add_argument("src_a",  help="First folder (A)")
    parser.add_argument("src_b",  help="Second folder (B)")
    parser.add_argument("output", nargs="?", default="comparison.xlsx",
                        help="Output file — .xlsx or .ods (default: comparison.xlsx)")
    parser.add_argument("--mode", default=None, choices=["quick","full"],
                        help="quick = no hash  |  full = hash included")
    parser.add_argument("--algo", default=None, choices=ALGORITHMS,
                        help="Skip benchmark and use this algorithm directly")
    parser.add_argument("--lang", default=None, choices=["en","tr"],
                        help="Language: en (default) or tr")
    args = parser.parse_args()

    for folder in (args.src_a, args.src_b):
        if not os.path.isdir(folder):
            # lang not yet known — print bilingual
            print(f"ERROR / HATA: '{folder}' is not a valid folder / geçerli klasör değil.")
            sys.exit(1)

    lang = lang_sec(args.lang)
    mode = args.mode if args.mode else mode_select(lang)
    algo = None
    if mode == "full":
        algo = args.algo if args.algo else benchmark_and_choose(lang)

    run(args.src_a, args.src_b, args.output, mode, algo, lang)

if __name__ == "__main__":
    main()
